from functools import wraps
from tempfile import NamedTemporaryFile
import re

from flask import Blueprint, redirect, url_for, render_template, flash, request, send_from_directory
from flask_login import current_user
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError
from sqlalchemy import or_
# from sqlalchemy_utils.types import Choice
# from werkzeug.utils import secure_filename

from .models import User, Class, Staff, Role
from .forms import StudentForm, TeacherForm, AdminForm, SearchUserForm, EditAdminForm
from .helper_func import process_data
from .. import db


def allowed_file(filename):
    pattern = r"^[\w\-]+?.(xlsx)$"
    check_file = re.search(pattern, filename)
    return '.' in filename and check_file is not None


# Activation needed. Move from here to dashboard folder
def activation_required(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if not current_user.has_activated:
            return redirect(url_for("reset_password"))

        return f(*args, **kwargs)

    return wrap


static_path = Path('.').parent.absolute() / 'modules/static'
users_bp = Blueprint("users", __name__, url_prefix="/users", static_folder=static_path)


# CRUD for class records

# View to show user registration page
@users_bp.route("/register")
def user_index():
    context = {}
    user_log = True
    admin = True  # remove this when user login is implemented
    student_form = StudentForm()
    teacher_form = TeacherForm()
    admin_form = AdminForm()
    context.update(admin=admin, student_form=student_form, user_log=user_log,
                   teacher_form=teacher_form, admin_form=admin_form)
    return render_template("users/add.html", **context)


# View to create student accounts
@users_bp.route("/register/user/student", methods=['POST'])
def add_student_account():
    role = Role.query.filter(Role.purpose == 'student').first()
    if role is None:
        return redirect(url_for(".user_index"))
    form = StudentForm()
    form.sid.data = str(form.sid.data).upper().replace("/", "_")
    form.name.data = str(form.name.data).lower()
    track, prog, year, _ = str(form.sid.data).upper().split("_", 3)
    student_user = User()
    form.populate_obj(student_user)

    if form.validate():
        print(form.validate_on_submit(), form.data)
        user_exist = User.query.filter(User.sid == form.sid.data).first()
        if user_exist is not None:
            flash(f"User with ID: {form.sid.data.replace('_', '/')} already exist.", "warning")
            return redirect(url_for(".user_index"))

        student_class = Class.query.filter(
            Class.programme == prog,
            Class.year_group == str(2000 + int(year)),
            Class.current_class == form.current_class.data,
            Class.track == track
        ).first()

        if student_class is None:
            flash("Class hasn't been created.", "info")
            return redirect(url_for(".user_index"))

        # Append user to role and class
        # TODO: student_user.password = form.password.data
        try:
            role.users.append(student_user)
            student_class.users.append(student_user)

            role.update()
            flash("User details added successfully.", "success")
        except IntegrityError:
            db.session.rollback()
            flash("User details not added.", "danger")

    return redirect(url_for(".user_index"))


# View to create teacher accounts
@users_bp.route("/register/user/teacher", methods=['POST'])
def add_teacher_account():
    role = Role.query.filter(Role.purpose == 'teacher').first()
    if role is None:
        return redirect(url_for(".user_index"))
    form = TeacherForm()
    form.sid.data = str(form.sid.data).upper().replace("/", "_")
    form.name.data = str(form.name.data).lower()
    teacher_user = User()
    form.populate_obj(teacher_user)
    if form.validate():
        print(form.validate_on_submit(), form.data)
        user_exist = User.query.filter(User.sid == form.sid.data).first()
        if user_exist is not None:
            flash(f"User with ID: {form.sid.data.replace('_', '/')} already exist.", "warning")
            return redirect(url_for(".user_index"))
        teacher_dept = form.department.data
        if teacher_dept is None:
            flash("Department hasn't been created.", "info")
            return redirect(url_for(".user_index"))
        # Append user to role and class
        # TODO: teacher_user.password = form.password.data
        try:
            role.users.append(teacher_user)
            teacher_dept.users.append(teacher_user)
            role.update()
            flash("User details added successfully.", "success")
        except IntegrityError:
            db.session.rollback()
            flash("User details not added.", "danger")
    return redirect(url_for(".user_index"))


# View to create admin accounts
@users_bp.route("/register/user/admin", methods=['POST'])
def add_admin_account():
    role = Role.query.filter(Role.purpose == 'admin').first()
    if role is None:
        return redirect(url_for(".user_index"))

    form = AdminForm()
    form.name.data = str(form.name.data).lower()
    admin_user = User()
    form.populate_obj(admin_user)
    if form.validate():
        user_exist = User.query.filter(User.sid == form.sid.data).first()
        if user_exist is not None:
            flash(f"User with ID: {form.sid.data} already exist.", "warning")
            return redirect(url_for(".user_index"))

        print(form.validate_on_submit(), form.data)
        # Append user to role and class
        # TODO: admin_user.password = form.password.data
        try:
            role.users.append(admin_user)
            role.update()
            flash("User details added successfully.", "success")
        except IntegrityError:
            db.session.rollback()
            flash("User details not added.", "danger")
    return redirect(url_for(".user_index"))


# View to show page for searching users
@users_bp.route("/list", methods=['GET', 'POST'])
def list_users():
    context = {}
    user_log = True
    admin = True  # remove this when user login is implemented
    search_form = SearchUserForm()

    if request.method == 'POST':
        search_keyword = str(search_form.search_term.data).replace('/', '_')
        get_accounts = User.query.filter(or_(User.sid.regexp_match(search_keyword.upper()),
                                             User.name.regexp_match(search_keyword.lower()))).all()
        context.update(user_records=get_accounts)
        return render_template("users/records_output.html", **context)

    context.update(admin=admin, search_form=search_form, user_log=user_log)
    return render_template("users/view.html", **context)


# View to Edit user records
@users_bp.route('/edit_user/<user_id>', methods=['GET', 'POST'])
def edit_user(user_id):
    context = {}
    # admin = True  # remove this when user login is implemented
    user_record = User.query.get(user_id)

    if user_record is None:
        return redirect(url_for(".list_users"))
    if request.method == 'POST':
        if user_record.role_id == 1:
            form = StudentForm(obj=user_record)
            form.sid.data = str(form.sid.data).upper().replace("/", "_")
            form.name.data = str(form.name.data).lower()
            track, prog, year, _ = str(form.sid.data).upper().split("_", 3)
            form.populate_obj(user_record)

            if form.validate():
                student_class = Class.query.filter(
                    Class.programme == prog,
                    Class.year_group == str(2000 + int(year)),
                    Class.current_class == form.current_class.data,
                    Class.track == track
                ).first()

                if student_class is None:
                    flash("Class hasn't been created.", "info")
                    return redirect(url_for(".user_index"))

                # Append user to role and class
                # TODO: student_user.password = form.password.data
                try:
                    student_class.update()
                    flash("User details edited successfully.", "success")
                except IntegrityError:
                    db.session.rollback()
                    flash("User details not edited.", "danger")

        if user_record.role_id == 2:
            form = TeacherForm(obj=user_record)
            if form.validate():
                teacher_dept = form.department.data
                form.sid.data = str(form.sid.data).upper().replace("/", "_")
                form.name.data = str(form.name.data).lower()
                form.populate_obj(user_record)
                if teacher_dept is None:
                    flash("Department hasn't been created.", "info")
                    return redirect(url_for(".user_index"))
                # Append user to role and class
                # TODO: teacher_user.password = form.password.data
                try:
                    teacher_dept.update()
                    flash("User details edited successfully.", "success")
                except IntegrityError:
                    db.session.rollback()
                    flash("User details not edited.", "danger")

        if user_record.role_id == 3:
            form = EditAdminForm(obj=user_record)
            form.populate_obj(user_record)

            print(form.validate_on_submit(), form.data)
            # Append user to role and class
            # TODO: admin_user.password = form.password.data
            try:
                user_record.update()
                flash("User details edited successfully.", "success")
            except IntegrityError:
                db.session.rollback()
                flash("User details not added.", "danger")

        return redirect(url_for(".list_users"))

    if user_record.role_id == 1:
        form = StudentForm(obj=user_record)
        form.sid.data = str(form.sid.data).upper().replace("_", "/")
        form.name.data = str(form.name.data).title()
    elif user_record.role_id == 2:
        form = TeacherForm(obj=user_record)
        form.sid.data = str(form.sid.data).upper().replace("_", "/")
        form.name.data = str(form.name.data).title()
        form.department.data = user_record.staff
    else:
        form = EditAdminForm(obj=user_record)
        form.name.data = str(form.name.data).title()
    context.update(form=form, user_id=user_id)
    return render_template("users/edit_record.html", **context)


# View to Delete user records
@users_bp.route('/delete_user/<user_id>', methods=['DELETE'])
def delete_user(user_id):
    # admin = True  # remove this when user login is implemented
    user_record = User.query.get(user_id)
    if user_record is not None:
        user_record.delete()
    msg = "Deleted user details successfully!"
    flash(msg, "success")
    context = {}
    # context.update(user_record=[])
    return render_template("users/records_output.html", **context)


# View to create users via file imports
@users_bp.route("/register/user/importusersfile", methods=['POST'])
def upload_users_file():
    if 'user_file' not in request.files:
        flash('No selected file', 'info')
        return redirect(url_for(".user_index"))
    file = request.files['user_file']
    if file.filename == '':
        flash('No selected file', 'info')
        return redirect(url_for(".user_index"))
    if file and allowed_file(file.filename):
        total_rows = request.form.get('total_rows', type=int)
        if total_rows is None:
            flash('Input the number of rows with data in file', 'warning')
            return redirect(url_for(".user_index"))
        # filename = secure_filename(file.filename)
        wb = load_workbook(file)

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)  # Save file in temporary file
            tmp.seek(0)

            wb2 = load_workbook(tmp)
            ws = wb2.active
            process_data(list(tuple(ws.iter_rows(
                max_col=6, min_row=2, max_row=total_rows, values_only=True)))
            )
        flash('Excel file imported successfully', 'success')
        return redirect(url_for(".user_index"))
    return redirect(url_for(".user_index"))

