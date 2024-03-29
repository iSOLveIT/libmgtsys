import re
from tempfile import NamedTemporaryFile

from flask import Blueprint, request, redirect, render_template, url_for, flash
from pathlib import Path
from sqlalchemy.exc import IntegrityError
from openpyxl import load_workbook

from project.modules.users.models import StudentClass, Role, Staff
from .forms import (
    AddClassForm,
    AddStaffForm,
    AddRoleForm,
    SearchStaffForm,
    SearchClassForm,
)
from ... import db, allowed_file
from .helper_func import process_data


static_path = Path(".").parent.absolute() / "modules/static"
record_mgt_bp = Blueprint(
    "record_mgt", __name__, url_prefix="/record_mgt", static_folder=static_path
)


# CRUD for class records

# View to show class page
@record_mgt_bp.route("/class")
def class_index():
    context = {}
    user_log = True
    admin = True  # remove this when user login is implemented
    form = AddClassForm()
    searchform = SearchClassForm()
    context.update(admin=admin, form=form, searchform=searchform, user_log=user_log)
    return render_template("records_mgt/class.html", **context)


# View to Create class records
@record_mgt_bp.route("/class/add_class", methods=["POST"])
def add_class():
    form = AddClassForm()

    if form.validate():
        print(form.validate_on_submit(), form.data)
        tag = f"{form.programme.data}{str(form.current_class.data).upper()}{form.track.data}{form.year_group.data}"
        class_exist = StudentClass.query.filter(StudentClass.class_tag == tag).first()
        if class_exist is not None:
            msg = "Class details already exist!"
            flash(msg, "warning")
            return redirect(url_for(".class_index"))
        try:
            student_class = StudentClass()
            student_class.class_tag = tag
            form.populate_obj(student_class)
            student_class.update()
            msg = "Created class details successfully"
            flash(msg, "success")
        except IntegrityError:
            # TODO: sqlalchemy.exc.IntegrityError: (psycopg2.errors.UniqueViolation) duplicate key value
            #  violates unique constraint "student_class_pkey". DETAIL:  Key (id)=(5) already exists.
            db.session.rollback()
            flash("Class details not added.", "danger")

    return redirect(url_for(".class_index"))


# View to Edit class records
@record_mgt_bp.route("/class/edit_class/<class_id>", methods=["GET", "POST"])
def edit_class(class_id):
    view = "class"
    context = {}
    admin = True  # remove this when user login is implemented
    class_record = StudentClass.query.get(class_id)
    if class_record is None:
        msg = "Class details not found!"
        flash(msg, "warning")
        return redirect(url_for(".class_index"))

    form = AddClassForm(obj=class_record)

    if request.method == "POST":
        if form.validate():
            print(form.validate_on_submit(), form.data)
            tag = f"{form.programme.data}{str(form.current_class.data).upper()}{form.track.data}{form.year_group.data}"
            class_record = (
                StudentClass.query.filter(StudentClass.id == class_id)
                .with_for_update()
                .first()
            )
            if class_record is None:
                msg = "Class details not found!"
                flash(msg, "warning")
                return redirect(url_for(".class_index"))
            try:
                form.populate_obj(class_record)
                class_record.class_tag = tag
                class_record.update()
                msg = "Updated class details successfully"
                flash(msg, "success")
            except IntegrityError:
                db.session.rollback()
                flash("Class details not edited.", "danger")

        return redirect(url_for(".class_index"))
    context.update(
        admin=admin, form=form, class_id=class_id, class_record=class_record, view=view
    )
    return render_template("records_mgt/edit_record.html", **context)


# View to Search class records
@record_mgt_bp.route("/class/search_class")
def search_class():
    view = "class"
    context = {}
    prog = request.args.get("programme", default=None)
    track = request.args.get("track", default=None)
    yr_grp = request.args.get("year_group")

    check_yr_grp = re.search(r"^\d{4}$", yr_grp)

    if check_yr_grp is not None:
        class_records = StudentClass.query.filter(
            StudentClass.programme == prog,
            StudentClass.track == track,
            StudentClass.year_group == yr_grp,
        ).all()
        context.update(class_records=class_records, view=view)
    return render_template("records_mgt/records_output.html", **context)


# View to Delete class records
@record_mgt_bp.route("/class/delete_class/<class_id>", methods=["DELETE"])
def delete_class(class_id):
    # admin = True  # remove this when user login is implemented
    class_record = StudentClass.query.get(class_id)
    if class_record is not None:
        class_record.delete()

    msg = "Deleted class details successfully!"
    return f"""
    <li class="breadcrumb-item fade" id="feedback">
        <svg xmlns="http://www.w3.org/2000/svg" style="display: none;">
          <symbol id="check-circle-fill" fill="currentColor" viewBox="0 0 16 16">
            <path d="M16 8A8 8 0 1 1 0 8a8 8 0 0 1 16 0zm-3.97-3.03a.75.75 0 0 0-1.08.022L7.477 9.417 5.384 7.323a.75.75 0 0 0-1.06 1.06L6.97 11.03a.75.75 0 0 0 1.079-.02l3.992-4.99a.75.75 0 0 0-.01-1.05z"/>
          </symbol>
        </svg>

        <span class="alert alert-success" role="alert">
            <svg width="30" height="20" role="img" aria-label="Success:"><use xlink:href="#check-circle-fill"/></svg>
            <span>{msg}</span>
        </span>
    </li>
    """


# View to create classes via file import
@record_mgt_bp.route("/class/add_class/importfile", methods=["POST"])
def upload_classes_file():
    if "classes_file" not in request.files:
        flash("No selected file", "info")
        return redirect(url_for(".class_index"))
    file = request.files["classes_file"]
    if file.filename == "":
        flash("No selected file", "info")
        return redirect(url_for(".class_index"))
    if file and allowed_file(file.filename):
        total_rows = request.form.get("total_rows", type=int)
        if total_rows is None:
            flash("Input the number of rows with data in file", "warning")
            return redirect(url_for(".class_index"))
        wb = load_workbook(file)

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)  # Save file in temporary file
            tmp.seek(0)

            wb2 = load_workbook(tmp)
            ws = wb2.active
            process_data(
                list(
                    tuple(
                        ws.iter_rows(
                            max_col=4, min_row=2, max_row=total_rows, values_only=True
                        )
                    )
                )
            )
        flash("Excel file imported successfully", "success")
        return redirect(url_for(".class_index"))
    return redirect(url_for(".class_index"))


# CRUD for staff records

# View to show staff page
@record_mgt_bp.route("/staff")
def staff_index():
    context = {}
    user_log = True
    admin = True  # remove this when user login is implemented
    form = AddStaffForm()
    searchform = SearchStaffForm()
    context.update(admin=admin, form=form, searchform=searchform, user_log=user_log)
    return render_template("records_mgt/staff.html", **context)


# View to Create staff records
@record_mgt_bp.route("/staff/add_staff", methods=["POST"])
def add_staff():
    form = AddStaffForm()
    staff_exist = Staff.query.filter(Staff.department == form.department.data).first()
    if staff_exist is not None:
        msg = "Staff details already exist!"
        flash(msg, "warning")
        return redirect(url_for(".class_index"))
    teacher_dept = Staff()
    form.department.data = str(form.department.data).capitalize()
    form.populate_obj(teacher_dept)

    if form.validate():
        print(form.validate_on_submit(), form.data)
        try:
            teacher_dept.update()
            msg = "Created staff details successfully"
            flash(msg, "success")
        except IntegrityError:
            msg = "Staff details already exist!"
            db.session.rollback()
            flash(msg, "danger")
    return redirect(url_for(".staff_index"))


# View to Edit staff records
@record_mgt_bp.route("/staff/edit_staff/<staff_id>", methods=["GET", "POST"])
def edit_staff(staff_id):
    view = "staff"
    context = {}
    admin = True  # remove this when user login is implemented
    staff_record = Staff.query.get(staff_id)
    form = AddStaffForm(obj=staff_record)
    form.populate_obj(staff_record)

    if request.method == "POST":
        if form.validate():
            print(form.validate_on_submit(), form.data)
            # prevent changes in selected row until you commit changes to DB
            staff_record = (
                Staff.query.filter(Staff.id == staff_id).with_for_update().first()
            )
            if staff_record is None:
                msg = "Staff details not found!"
                flash(msg, "warning")
                return redirect(url_for(".staff_index"))
            try:
                staff_record.department = str(staff_record.department).capitalize()
                staff_record.update()
                msg = "Updated staff details successfully"
                flash(msg, "success")
            except IntegrityError:
                msg = "Staff details already exist!"
                db.session.rollback()
                flash(msg, "danger")
        return redirect(url_for(".staff_index"))
    context.update(
        admin=admin, form=form, staff_id=staff_id, staff_record=staff_record, view=view
    )
    return render_template("records_mgt/edit_record.html", **context)


# View to Search staff records
@record_mgt_bp.route("/staff/search_staff")
def search_staff():
    view = "staff"
    context = {}
    dept = int(request.args.get("department", default=None))

    staff_record = Staff.query.filter(Staff.id == dept).first()
    if staff_record is not None:
        context.update(dept_record=staff_record)
    context.update(view=view)
    return render_template("records_mgt/records_output.html", **context)


# View to Delete staff records
@record_mgt_bp.route("/staff/delete_staff/<staff_id>", methods=["DELETE"])
def delete_staff(staff_id):
    # admin = True  # remove this when user login is implemented
    staff_record = Staff.query.get(staff_id)
    if staff_record is not None:
        staff_record.delete()

    msg = "Deleted staff details successfully"
    return f"""
    <li class="breadcrumb-item fade" id="feedback">
        <svg xmlns="http://www.w3.org/2000/svg" style="display: none;">
          <symbol id="check-circle-fill" fill="currentColor" viewBox="0 0 16 16">
            <path d="M16 8A8 8 0 1 1 0 8a8 8 0 0 1 16 0zm-3.97-3.03a.75.75 0 0 0-1.08.022L7.477 9.417 5.384 7.323a.75.75 0 0 0-1.06 1.06L6.97 11.03a.75.75 0 0 0 1.079-.02l3.992-4.99a.75.75 0 0 0-.01-1.05z"/>
          </symbol>
        </svg>

        <span class="alert alert-success" role="alert">
            <svg width="30" height="20" role="img" aria-label="Success:"><use xlink:href="#check-circle-fill"/></svg>
            <span>{msg}</span>
        </span>
    </li>
    """


# CRUD for roles records

# View to show roles page
@record_mgt_bp.route("/role")
def role_index():
    context = {}
    admin = True  # remove this when user login is implemented
    user_log = True
    form = AddRoleForm()
    role_records = Role.query.all()
    context.update(admin=admin, form=form, role_records=role_records, user_log=user_log)
    return render_template("records_mgt/role.html", **context)


# View to Create role records
@record_mgt_bp.route("/role/add_role", methods=["POST"])
def add_role():
    view = "role"
    context = {}
    user_log = True
    form = AddRoleForm()
    account_type = Role()
    form.populate_obj(account_type)
    if form.validate():
        account_type.permission_level = True if form.purpose.data == "admin" else False
        try:
            account_type.update()
        except IntegrityError:
            db.session.rollback()

    role_records = Role.query.all()
    context.update(view=view, user_log=user_log, role_records=role_records)
    return render_template("records_mgt/records_output.html", **context)


# View to Edit role records
@record_mgt_bp.route("/role/edit_role/<role_id>", methods=["GET", "POST"])
def edit_role(role_id):
    view = "role"
    context = {}
    admin = True  # remove this when user login is implemented
    role_record = Role.query.get(role_id)
    form = AddRoleForm(obj=role_record)
    form.populate_obj(role_record)
    # print(form.validate_on_submit(), form.data)
    if request.method == "POST":
        if form.validate():
            print(form.validate(), form.data)
            if role_record is None:
                msg = "Role does not exist!"
                flash(msg, "danger")
                return redirect(url_for(".role_index"))
            role_record.permission_level = (
                True if form.purpose.data == "admin" else False
            )
            try:
                role_record.update()
            except IntegrityError:
                db.session.rollback()

        return redirect(url_for(".role_index"))
    context.update(
        admin=admin, form=form, role_id=role_id, role_record=role_record, view=view
    )
    return render_template("records_mgt/edit_record.html", **context)


# View to Delete role records
@record_mgt_bp.route("/role/delete_role/<role_id>", methods=["DELETE"])
def delete_role(role_id):
    # admin = True  # remove this when user login is implemented
    role_record = Role.query.get(role_id)
    if role_record is not None:
        role_record.delete()

    msg = "Deleted role successfully"
    return f"""
    <li class="breadcrumb-item fade" id="feedback">
        <svg xmlns="http://www.w3.org/2000/svg" style="display: none;">
          <symbol id="check-circle-fill" fill="currentColor" viewBox="0 0 16 16">
            <path d="M16 8A8 8 0 1 1 0 8a8 8 0 0 1 16 0zm-3.97-3.03a.75.75 0 0 0-1.08.022L7.477 9.417 5.384 7.323a.75.75 0 0 0-1.06 1.06L6.97 11.03a.75.75 0 0 0 1.079-.02l3.992-4.99a.75.75 0 0 0-.01-1.05z"/>
          </symbol>
        </svg>

        <span class="alert alert-success" role="alert">
            <svg width="30" height="20" role="img" aria-label="Success:"><use xlink:href="#check-circle-fill"/></svg>
            <span>{msg}</span>
        </span>
    </li>
    """
