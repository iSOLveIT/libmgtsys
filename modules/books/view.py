from flask import Blueprint, render_template, flash, redirect, url_for, request
from pathlib import Path
from tempfile import NamedTemporaryFile

from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from openpyxl import load_workbook

from .models import Books
from .forms import AddBooksForm, SearchBooksForm
from .. import db, allowed_file
from .helper_func import process_data

static_path = Path(".").parent.absolute() / "modules/static"
books_bp = Blueprint("books", __name__, url_prefix="/books", static_folder=static_path)


# TODO: Remove the book table since it is unnecessary. Instead search for a book using
#  the classification number or the title or author or category
#  and their added as borrowed_users to that book.
#  We then create the user_book table using that the book and user entities.


# View to show book inventory page
@books_bp.route("/add")
def book_index():
    context = {}
    user_log = True
    admin = True  # remove this when user login is implemented
    books_form = AddBooksForm()
    context.update(admin=admin, books_form=books_form, user_log=user_log)
    return render_template("books/add.html", **context)


# View to show page for searching books
@books_bp.route("/list", methods=["GET", "POST"])
def list_books():
    context = {}
    user_log = True
    admin = True  # remove this when user login is implemented
    search_form = SearchBooksForm()

    if request.method == "POST":
        search_keyword = str(search_form.search_term.data).lower()
        get_books = []
        if len(search_keyword) != 0:
            get_books = Books.query.filter(
                or_(
                    Books.title.regexp_match(search_keyword),
                    Books.author.regexp_match(search_keyword),
                    Books.classification_no.regexp_match(search_keyword),
                    Books.category.regexp_match(search_keyword),
                )
            ).all()
        context.update(book_records=get_books)
        return render_template("books/records_output.html", **context)

    context.update(admin=admin, search_form=search_form, user_log=user_log)
    return render_template("books/view.html", **context)


# View to add books
@books_bp.route("/add/books", methods=["POST"])
def add_books():
    form = AddBooksForm()

    if form.validate():
        print(form.validate_on_submit(), form.data)
        books = Books()
        form.title.data = str(form.title.data).lower()
        form.classification_no.data = str(form.classification_no.data).lower()
        form.author.data = str(form.author.data).lower()
        form.publication.data = str(form.publication.data).lower()

        book_exist = Books.query.filter(
            Books.classification_no == form.classification_no.data
        ).first()
        if book_exist is not None:
            book_exist.current_qty = book_exist.current_qty + int(form.qty_added.data)
            form.populate_obj(book_exist)
            try:
                book_exist.update()
                flash("Books detail already exist so quantity was updated.", "success")
            except IntegrityError:
                db.session.rollback()
                flash("Books detail not added.", "danger")
        else:
            form.populate_obj(books)
            last_book = Books.query.order_by(Books.id.desc()).first()
            access_no = last_book.access_no + int(form.qty_added.data)
            books.qty_spoilt, books.current_qty, books.access_no = (
                0,
                int(form.qty_added.data),
                access_no,
            )

            try:
                books.update()
                flash("Books detail added successfully.", "success")
            except IntegrityError:
                db.session.rollback()
                flash("Books detail not added.", "danger")

    return redirect(url_for(".book_index"))


# View to Edit book records
@books_bp.route("/edit_book/<book_id>", methods=["GET", "POST"])
def edit_book(book_id):
    context = {}
    # admin = True  # remove this when user login is implemented
    book_record = Books.query.get(book_id)

    if book_record is None:
        return redirect(url_for(".list_books"))
    if request.method == "POST":
        form = AddBooksForm()
        if form.validate():
            form.title.data = str(form.title.data).lower()
            form.classification_no.data = str(form.classification_no.data).lower()
            form.author.data = str(form.author.data).lower()
            form.publication.data = str(form.publication.data).lower()

            book_record.current_qty = book_record.current_qty - book_record.qty_added
            book_record.current_qty = book_record.current_qty + int(form.qty_added.data)

            form.populate_obj(book_record)

            try:
                book_record.update()
                flash("Book details edited successfully.", "success")
            except IntegrityError:
                db.session.rollback()
                flash("Books details not edited.", "danger")
        return redirect(url_for(".list_books"))

    form = AddBooksForm(obj=book_record)
    form.title.data = str(form.title.data).upper()
    form.classification_no.data = str(form.classification_no.data).upper()
    form.author.data = str(form.author.data).title()
    form.publication.data = str(form.publication.data).title()

    context.update(form=form, book_id=book_id)
    return render_template("books/edit_record.html", **context)


# View to Delete book records
@books_bp.route("/delete_book/<book_id>", methods=["DELETE"])
def delete_book(book_id):
    # admin = True  # remove this when user login is implemented
    book_record = Books.query.get(book_id)
    if book_record is not None:
        book_record.delete()
    msg = "Deleted book details successfully!"

    return f"""
<li class="breadcrumb-item disappear" id="feedback">
    <svg xmlns="http://www.w3.org/2000/svg" style="display: none;">
      <symbol id="check-circle-fill" fill="currentColor" viewBox="0 0 16 16">
        <path d="M16 8A8 8 0 1 1 0 8a8 8 0 0 1 16 0zm-3.97-3.03a.75.75 0 0 0-1.08.022L7.477 9.417 5.384 7.323a.75.75 0 0 0-1.06 1.06L6.97 11.03a.75.75 0 0 0 1.079-.02l3.992-4.99a.75.75 0 0 0-.01-1.05z"/>
      </symbol>
    </svg>
        
    <span class="alert alert-success" role="alert">
        <svg width="30" height="20" role="img" aria-label="Success:"><use xlink:href="#check-circle-fill"/></svg>
        <span>{msg}</span>
    </span>
</li>
"""


# View to add books via file import
@books_bp.route("/add/books/importfile", methods=["POST"])
def upload_books_file():
    if "books_file" not in request.files:
        flash("No selected file", "info")
        return redirect(url_for(".book_index"))
    file = request.files["books_file"]
    if file.filename == "":
        flash("No selected file", "info")
        return redirect(url_for(".book_index"))
    if file and allowed_file(file.filename):
        total_rows = request.form.get("total_rows", type=int)
        if total_rows is None:
            flash("Input the number of rows with data in file", "warning")
            return redirect(url_for(".book_index"))
        wb = load_workbook(file)

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)  # Save file in temporary file
            tmp.seek(0)

            wb2 = load_workbook(tmp)
            ws = wb2.active
            process_data(
                list(
                    tuple(
                        ws.iter_rows(
                            max_col=8, min_row=2, max_row=total_rows, values_only=True
                        )
                    )
                )
            )
        flash("Excel file imported successfully", "success")
        return redirect(url_for(".book_index"))
    return redirect(url_for(".book_index"))
