from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
import re
from flask import Flask, flash, request, redirect, render_template
from werkzeug.utils import secure_filename


app = Flask(__name__)
app.config["SECRET_KEY"] = "UPLOAD_FOLDER"
app.config["MAX_CONTENT_LENGTH"] = 2 * 1000 * 1000


def allowed_file(filename):
    pattern = r"^[\w\-]+?.(xlsx)$"
    check_file = re.search(pattern, filename)
    return "." in filename and check_file is not None


def process_data(header, file_data):
    return dict(
        tbl_head=header, tbl_content=sorted(list(set(file_data)), key=lambda x: x[1])
    )


@app.route("/", methods=["GET", "POST"])
def upload_file():
    if request.method == "POST":
        # check if the post request has the file part
        if "file" not in request.files:
            flash("No file part")
            return redirect(request.url)
        file = request.files["file"]
        # If the user does not select a file, the browser submits an
        # empty file without a filename.
        if file.filename == "":
            flash("No selected file")
            return redirect(request.url)
        if file and allowed_file(file.filename):
            total_rows = int(request.form.get("total_rows"))
            filename = secure_filename(file.filename)
            print(filename)
            wb = load_workbook(file)

            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)  # Save file in temporary file
                tmp.seek(0)

                wb2 = load_workbook(tmp)
                ws = wb2.active
                # TODO: Remove the first row as header from the list of rows
                #  then send it as a parameter with the new list
                # get_data = list(tuple(ws.iter_rows(max_col=6, min_row=1, max_row=1382, values_only=True)))
                # print(len(tuple(ws.iter_rows())))
                data_header = list(
                    tuple(ws.iter_rows(max_col=6, max_row=1, values_only=True))
                ).pop(0)
                content = process_data(
                    data_header,
                    list(
                        tuple(
                            ws.iter_rows(
                                max_col=6,
                                min_row=2,
                                max_row=total_rows,
                                values_only=True,
                            )
                        )
                    ),
                )

            return render_template("upload_file.html", content=content)

    return render_template("upload_file.html")


if __name__ == "__main__":
    app.run()
